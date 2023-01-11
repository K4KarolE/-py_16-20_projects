# randomly picking a movie from MoviePY excel sheet
# give a hint about the movie (release year, stars, director,..)
# asking the user to guess the title

import random
from openpyxl import Workbook, load_workbook
wb = load_workbook('MoviePY.xlsx', data_only=True)
ws = wb.active

#data_only=True -> copying values from excel instead of formulas for the Haveseen cell value

#GAME - BANNER
print()
k = 50
print('*'*k)    
print(' WHICH MOVIE BRO? '.center(k,'*'))
print('*'*k)

round = 0
score = 0

while True:
# VALUE EXTRACTION FROM THE EXCEL
    cellnumber = random.randrange(6,6951)

    cell = 'C' + str(cellnumber)
    movietitle = ws[cell].value

    while movietitle == None:        # If it picked a empty cell(1 title in at least 3 merged cells), 
        cellnumber += 1              # looking for the next valid non-empty one
        cell = 'C' + str(cellnumber)
        movietitle = ws[cell].value
    
    if movietitle == '-':           # some hungarian movies do not have english title
        cell = 'D' + str(cellnumber)
        movietitle = ws[cell].value

    cellRYear = 'E' + str(cellnumber)
    ReleaseYear = ws[cellRYear].value

    cellHSeen = 'N' + str(cellnumber)
    HaveSeen = ws[cellHSeen].value

    cellDirector_1 = 'F' + str(cellnumber)
    Director_1 = ws[cellDirector_1].value
    cellDirector_2 = 'F' + str(cellnumber + 1)
    Director_2 = ws[cellDirector_2].value
    cellDirector_3 = 'F' + str(cellnumber + 2)
    Director_3 = ws[cellDirector_3].value

    cellStar_1 = 'G' + str(cellnumber)
    Star_1 = ws[cellStar_1].value
    cellStar_2 = 'G' + str(cellnumber + 1)
    Star_2 = ws[cellStar_2].value
    cellStar_3 = 'G' + str(cellnumber + 2)
    Star_3 = ws[cellStar_3].value


# - THE GAME -
#ROUND - BANNER
    round +=1
    print()
    print((' ROUND ' + str(round) + ' ').center(50, '-'))
    print
    print('It was released in ' + str(ReleaseYear) + '.')
    print()

# MOVIE INFO/HINT
    if Director_1 != None:
        print('Director(s): ')
        print(str(Director_1).strip().rjust(len('Director(s):' + str(Director_1).strip())))
        if Director_2 != None:
            print(str(Director_2).strip().rjust(len('Director(s):' + str(Director_2).strip())))
        if Director_3 != None:
            print(str(Director_3).strip().rjust(len('Director(s):' + str(Director_3).strip())))

    if Star_1 != None:
        print('Star(s): ')
        print(str(Star_1).strip().rjust(len('Star(s):' + str(Star_1).strip())))
        if Star_2 != None:
            print(str(Star_2).strip().rjust(len('Star(s):' + str(Star_2).strip())))
        if Star_3 != None:
            print(str(Star_3).strip().rjust(len('Star(s):' + str(Star_3).strip())))

    print()
    if HaveSeen == 1: 
        print('You have seen this movie only once since 05/2012.')
    else:
        print('You have seen this movie ' + str(HaveSeen) +  ' times since 05/2012.')
    print('-'*k + '\n')

#USER ANSWER
    print('Type:\n"R" to reveal the movie title\n"Q" for leave the game')
    answer = input('Your take: ').lower()
    
    if answer == 'q':
        print()
        print('Thx for playing!')
        print()
        break

    if answer == 'r':
        print()
        print('The movie was: ' + str(movietitle).strip())

    print()
    bePart = 0  # variable for a partially correct answers

    if answer == str(movietitle).lower():
        print('That is correct!')
        score += 1
    else:
        for i in answer.split():                # creating a list from the answer, checking the items in the movie title
            if i in str(movietitle).lower().split():
                bePart += 1
        if bePart > len((str(movietitle).split())) * 0.5:      # answer matching by 50% of he movie title  - scenario
            print('Almost! The movie was: ' + str(movietitle))
            score += 0.5
        elif answer != 'r':
            print('Incorrect! The movie was: ' + str(movietitle))

    print('Your score: ' + str(score) + '/' + str(round))
    print()
    print('Type:\n"Q" for leave the game\n"N" for the next round')
    answer = input('Next step: ').lower()
    while answer not in ['q', 'n']:
        print()
        print('Please try again.')
        answer = input('Next step: ').lower()
    
    if answer == 'q':
        print()
        print('Thx for playing!')
        print()
        break
    else:
        continue
    