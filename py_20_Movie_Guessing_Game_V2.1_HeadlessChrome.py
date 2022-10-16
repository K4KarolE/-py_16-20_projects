''' Movie Guessing Game V2 - Excel + Selenium
- pick a random title + it`s release year from the MoviePY excel sheet
- look for it in a search engine combined with a movie database title / using search engine as a pre-step,
  because (apart from IMDb) not able to search by movie title + release year, which makes our search more precise 
- open the first match = open the title`s movie database site
- take the synopsys/plot of the movie and display it for the user
- ask the user to guess the title
- give a hint/help(director, release year, stars..) if the user needed '''

import random
from openpyxl import Workbook, load_workbook
from requests import options
wb = load_workbook('MoviePY.xlsx', data_only=True)
ws = wb.active

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

options = webdriver.ChromeOptions()
options.headless = True

#GAME - BANNER
print()
k = 50
print('*'*k)    
print(' Please wait, the program is loading '.center(k,'*'))
print('*'*k)

round = 0
score = 0

while True:
# VALUE EXTRACTION FROM THE EXCEL
    cellnumber = random.randrange(6,6948)

    cell = 'C' + str(cellnumber)
    movietitle = ws[cell].value

    cellRYear = 'E' + str(cellnumber)
    releaseYear = ws[cellRYear].value

    def isItShow():                           # movie titles including S01, S02.. or it`s release year like "1992-1999" are TV shows
        if len(str(releaseYear)) != 4:
            return True
        else:
            for i in str(movietitle).split():
                if i[0] == 'S' and len(i) == 3 and i[1].isdigit() and i[2].isdigit():
                    return True

    while movietitle == None or movietitle == '-' or isItShow():        # movietitle == None: empty cell(1 title in at least 3 merged cells) 
        cellnumber += 1                                                 # movietitle == '-': excluding movies with no english title
        cell = 'C' + str(cellnumber)                                    # isItShow: excluding TV shows
        movietitle = ws[cell].value

    cellRYear = 'E' + str(cellnumber)
    releaseYear = ws[cellRYear].value

    cellHSeen = 'N' + str(cellnumber)
    haveSeen = ws[cellHSeen].value

# VALUE EXTRACTION FROM WEBSITE
    try:
        PATH = 'C:\Program Files (x86)\chromedriver.exe'
        driver = webdriver.Chrome(PATH, chrome_options=options)
        driver.get('https://duckduckgo.com/')  # using search engine as a pre step
                                            # because apart from IMDb, not able to search by movie title + release year
        try:
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.NAME, 'q' ))
            )
        except:
            driver.quit()
        
        movieDataBase = 'rotten tomatoes'   # the movie data base / unfortunately I was not able to use IMDb
        searchForMovie = ' '.join([ movieDataBase, str(movietitle), str(releaseYear) ])

# LOOKING FOR THE MOVIE VIA SEARCH ENGINE
        search = driver.find_element(By.NAME,'q')
        search.send_keys(searchForMovie)
        search.send_keys(Keys.RETURN)

# CLICKING ON THE FIRST RESULT
        search = driver.find_element(By.ID,'r1-0')
        search.click()

# COLLECTING THE INFORMATION FOR THE MOVIE FROM THE MOVIE DATABASE
        plot = driver.find_element(By.ID, 'movieSynopsis').text
        director = driver.find_element(By.CSS_SELECTOR, 'li.meta-row:nth-child(4) > div:nth-child(2) > a:nth-child(1)').text
        star_1 = driver.find_element(By.CSS_SELECTOR, 'div.cast-item:nth-child(1) > div:nth-child(2) > a:nth-child(1) > span:nth-child(1)').text
        star_2 = driver.find_element(By.CSS_SELECTOR, 'div.cast-item:nth-child(2) > div:nth-child(2) > a:nth-child(1) > span:nth-child(1)').text
        star_3 = driver.find_element(By.CSS_SELECTOR, 'div.cast-item:nth-child(3) > div:nth-child(2) > a:nth-child(1) > span:nth-child(1)').text
        driver.quit()

# - THE GAME -
# ROUND - BANNER
        round +=1
        print('\n')
        print((' ROUND ' + str(round) + ' ').center(50, '-'))
        print()
        print('Guess the movie where..')
        print(plot)
        print()

# USER ANSWER
        print('Type:\n"M" for more information\n"R" to reveal the movie title\n"Q" for leave the game')
        answer = input('Your take: ').lower()
        
        if answer == 'q':
            print()
            print('Thx for playing!')
            print()
            break

        if answer == 'r':
            print()
            print('The movie was: ' + str(movietitle).strip())

        if answer == 'm':
            print()
            print('A little help:')
            print('Year of release: ' + str(releaseYear))
            print('Director: ' + director)
            print('Star(s): ' + star_1 + ', ' + star_2 + ', ' + star_3)
            if haveSeen == 1: 
                print('You have seen this movie only once since 05/2012.')
            else:
                print('You have seen this movie ' + str(haveSeen) +  ' times since 05/2012.')
            print()
            answer = input('Your take: ').lower()

        print()
        bePart = 0  # variable for a partially correct answers

        if answer == str(movietitle).lower():
            print('That is correct!')
            score += 1
        else:
            for i in answer.split():                # creating a list from the answer, checking the items in the movie title
                if i in str(movietitle).lower().split():
                    bePart += 1
            if bePart >= len((str(movietitle).split())) * 0.5:      # answer matching by >= 50% of he movie title - scenario
                print('Almost! The movie was: ' + str(movietitle))  # of course it is not ideal, when the title contains only two parts and one of them is "the"
                score += 0.5                                        # we are going leave it like this for now, the main objective was the excel and selenium usage
            elif answer != 'r':
                print('Incorrect! The movie was: ' + str(movietitle))

# AFTER ROUND - USER INPUT
        print('Your score: ' + str(score) + '/' + str(round))
        print()
        print('Type:\n"Q" for leave the game\n"N" for the next round')
        answer = input('Next step: ').lower()
        while answer not in ['q', 'n']:
            print()
            print('Please try again.')
            answer = input('Next step: ').lower()
        
        if answer == 'q':
            print()
            print('Thx for playing!')
            print()
            break
        else:
            continue
    except:             # the search engine - first result - movie database combination is not ideal
        print()         # if there is an error in this sequence
        k = 50          # below message is displayed and looking for a new title from the excel/movie database
        print('*'*k)    
        print(' Looking for a new title '.center(k,'*'))
        print('*'*k)